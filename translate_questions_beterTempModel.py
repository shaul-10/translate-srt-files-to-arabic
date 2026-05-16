"""
translate_questions.py
----------------------
Translates a JSON file of questions from Hebrew to Arabic using OpenAI API.
Generates an Arabic JSON file + Excel review file.

Usage:
    python3 translate_questions.py while_level01.json

Requirements:
    pip install openai openpyxl

Set API Key - option 1 (environment variable, recommended):
    export OPENAI_API_KEY="sk-..."

Set API Key - option 2 (directly in code, line 25):
    API_KEY = "sk-..."
"""

import json
import sys
import os
from pathlib import Path

from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── API Key ─────────────────────────────────────────────────────
API_KEY = os.environ.get("OPENAI_API_KEY", "")
# If no environment variable, insert key here:
# API_KEY = "sk-..."
# ─────────────────────────────────────────────────────────────────

MODEL = "gpt-4o"

# ─── Technical Glossary (Hebrew → Arabic) ─────────────────────
GLOSSARY = {
    "לולאת while": "حلقة while",
    "שיטה סטטית": "دالة ساكنة",
    "מספר שלם": "عدد صحيح",
    "ספרה": "رقم",
    "סכום": "مجموع",
    "הדפסה": "طباعة",
    "מחזיר": "يُرجع",
    "מקבל": "يستقبل",
    "אלגוריתם": "خوارزمية",
    "אינדקס": "فهرس",
    "מערך": "مصفوفة",
    "משתנה": "متغير",
    "לולאה": "حلقة",
    "תנאי": "شرط",
    "פונקציה": "دالة",
}


def translate_question(client, question: dict) -> dict:
    """Translates a single question from Hebrew to Arabic using technical terminology."""
    prompt = f"""You are a professional technical translator from Hebrew to Arabic,
specializing in programming education for high school students.

TECHNICAL TRANSLATION RULES:
- Use standard computer science terminology in Arabic
- Maintain consistency with programming vocabulary
- Use formal, technical Arabic suitable for developers
- Each term should be precise and unambiguous
- Preserve all variable names, method names, and English technical terms
- Use the following glossary for technical terms:
{json.dumps(GLOSSARY, ensure_ascii=False, indent=2)}

Translate the following JSON question object from Hebrew to Arabic.

STRICT RULES:
- Keep all fields with type "code" UNCHANGED — do not translate code
- Translate: topic, text, options (array items), explanation items with type "he"
- Keep the exact same JSON structure
- Arabic must be natural and appropriate for high school students
- Return ONLY valid JSON, no markdown, no explanation

Input:
{json.dumps(question, ensure_ascii=False, indent=2)}"""

    response = client.chat.completions.create(
        model=MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,  # Low temperature for consistent, technical translation
    )

    raw = response.choices[0].message.content.strip()
    # Remove markdown if present
    raw = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def option_to_string(option):
    """Converts an option (string or object) to a string."""
    if isinstance(option, str):
        return option
    elif isinstance(option, dict) and "text" in option:
        return option["text"]
    else:
        return str(option)


def save_excel(he_questions: list, ar_questions: list, output_path: str):
    """Saves an Excel file with Hebrew, Arabic and a correction column."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translation"

    headers = [
        "Question #", "Topic (Hebrew)", "Topic (Arabic)",
        "Text (Hebrew)", "Text (Arabic)",
        "Correct Answer (Hebrew)", "Correct Answer (Arabic)",
        "All Options (Hebrew)", "All Options (Arabic)",
        "Correction"
    ]
    col_widths = [12, 18, 18, 30, 30, 20, 20, 35, 35, 30]

    header_fill = PatternFill("solid", fgColor="EEEDFE")
    fix_fill    = PatternFill("solid", fgColor="FFF3CD")
    header_font = Font(bold=True, color="3C3489")
    fix_font    = Font(bold=True, color="856404")

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fix_fill if h == "Correction" else header_fill
        cell.font = fix_font if h == "Correction" else header_font
        cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20

    for i, (hq, aq) in enumerate(zip(he_questions, ar_questions), 2):
        he_ans = hq.get("options", [""])[hq.get("answer", 0)]
        ar_ans = aq.get("options", [""])[aq.get("answer", 0)]
        
        # המרת options ל-strings לפני join
        he_options_str = " | ".join([option_to_string(opt) for opt in hq.get("options", [])])
        ar_options_str = " | ".join([option_to_string(opt) for opt in aq.get("options", [])])

        row = [
            i - 1,
            hq.get("topic", ""),
            aq.get("topic", ""),
            hq.get("text", ""),
            aq.get("text", ""),
            option_to_string(he_ans),
            option_to_string(ar_ans),
            he_options_str,
            ar_options_str,
            ""  # Correction column — empty
        ]

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

        ws.row_dimensions[i].height = 40

    ws.freeze_panes = "A2"
    wb.save(output_path)


def print_help():
    print("""
translate_questions.py — Hebrew to Arabic question translator
=============================================================

USAGE:
    python3 translate_questions.py <json_file>

ARGUMENTS:
    <json_file>   Path to a JSON file containing Hebrew questions
                  (e.g. while_level01.json)

OUTPUT (saved in the same folder as the input file):
    <name>_ar.json       Arabic JSON — ready to use in the quiz builder
    <name>_review.xlsx   Excel file for review, with an empty "Correction" column

SETUP (one time only):
    1. Activate virtual environment:
           source venv/bin/activate
    2. Install dependencies:
           pip install openai openpyxl
    3. Set your OpenAI API key:
           export OPENAI_API_KEY="sk-..."

EXAMPLES:
    python3 translate_questions.py while_level01.json
    python3 translate_questions.py ../questions/for_level02.json

OPTIONS:
    -h, --help, -?    Show this help message
""")


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help", "-?"):
        print_help()
        sys.exit(0)

    if not API_KEY:
        print("Error: OPENAI_API_KEY is not set.")
        print("Run: export OPENAI_API_KEY='sk-...'")
        print("Or insert the key directly in line 25 of the script.")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Error: file '{input_path}' not found.")
        sys.exit(1)

    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        he_questions = data
    else:
        he_questions = data.get("questionsHe") or data.get("questions") or []
    if not isinstance(he_questions, list) or not he_questions:
        print("Error: no questions found in file.")
        sys.exit(1)

    print(f"Loaded {len(he_questions)} questions from {input_path.name}")

    client = OpenAI(api_key=API_KEY)
    ar_questions = []

    for i, q in enumerate(he_questions, 1):
        print(f"Translating question {i}/{len(he_questions)}...", end=" ", flush=True)
        try:
            ar_q = translate_question(client, q)
            ar_questions.append(ar_q)
            print("done")
        except Exception as e:
            print(f"ERROR: {e}")
            sys.exit(1)

    base = input_path.stem
    json_out  = input_path.parent / f"{base}_ar.json"
    excel_out = input_path.parent / f"{base}_review.xlsx"

    with open(json_out, "w", encoding="utf-8") as f:
        json.dump({"questionsAr": ar_questions}, f, ensure_ascii=False, indent=2)

    save_excel(he_questions, ar_questions, str(excel_out))

    print(f"\nDone!")
    print(f"  Arabic JSON : {json_out}")
    print(f"  Excel review: {excel_out}")


if __name__ == "__main__":
    main()
